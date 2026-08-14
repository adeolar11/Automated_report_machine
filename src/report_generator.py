import os

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from .config import REPORT_DIR
from .analyzer import calculate_metrics


def create_excel_report(
    df,
    chart_path
):

    os.makedirs(
        REPORT_DIR,
        exist_ok=True
    )

    metrics = calculate_metrics(df)

    wb = Workbook()


    # ---------------------------------
    # STYLES
    # ---------------------------------

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_fill = PatternFill(
        start_color="2C3E50",
        end_color="2C3E50",
        fill_type="solid"
    )

    center_align = Alignment(
        horizontal="center"
    )


    # ---------------------------------
    # EXECUTIVE SUMMARY
    # ---------------------------------

    ws1 = wb.active

    ws1.title = "Executive Summary"

    ws1.merge_cells("A1:F1")

    ws1["A1"] = (
        "WEEKLY SALES PERFORMANCE REPORT"
    )

    ws1["A1"].font = Font(
        size=18,
        bold=True
    )

    ws1["A1"].alignment = center_align


    kpis = [
        ["Metric", "Value"],
        [
            "Total Revenue",
            metrics["total_revenue"]
        ],
        [
            "Total Orders",
            metrics["total_orders"]
        ],
        [
            "Total Quantity",
            metrics["total_quantity"]
        ],
        [
            "Average Order Value",
            metrics["avg_order_value"]
        ],
        [
            "Top Region",
            metrics["top_region"]
        ],
        [
            "Top Product",
            metrics["top_product"]
        ],
        [
            "Holiday Revenue",
            metrics["holiday_revenue"]
        ],
        [
            "Non-Holiday Revenue",
            metrics["non_holiday_revenue"]
        ]
    ]


    for row_number, row in enumerate(
        kpis,
        start=3
    ):

        for column_number, value in enumerate(
            row,
            start=1
        ):

            cell = ws1.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            if row_number == 3:

                cell.font = header_font

                cell.fill = header_fill

                cell.alignment = center_align


    # ---------------------------------
    # RAW DATA
    # ---------------------------------

    ws2 = wb.create_sheet(
        "Raw Data"
    )

    for col_idx, col in enumerate(
        df.columns,
        start=1
    ):

        cell = ws2.cell(
            row=1,
            column=col_idx,
            value=col
        )

        cell.font = header_font

        cell.fill = header_fill


    for row_idx, row in enumerate(
        df.itertuples(index=False),
        start=2
    ):

        for col_idx, value in enumerate(
            row,
            start=1
        ):

            ws2.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )


    # ---------------------------------
    # REGION SALES
    # ---------------------------------

    ws3 = wb.create_sheet(
        "Region Sales"
    )

    region_sales = (
        df.groupby("Country")
        .agg(
            Revenue=("Revenue", "sum"),
            Quantity=("Quantity", "sum")
        )
        .sort_values(
            "Revenue",
            ascending=False
        )
        .reset_index()
    )


    for col_idx, col in enumerate(
        region_sales.columns,
        start=1
    ):

        cell = ws3.cell(
            row=1,
            column=col_idx,
            value=col
        )

        cell.font = header_font

        cell.fill = header_fill


    for row_idx, row in enumerate(
        region_sales.itertuples(
            index=False
        ),
        start=2
    ):

        for col_idx, value in enumerate(
            row,
            start=1
        ):

            ws3.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )


    # ---------------------------------
    # ADD CHART
    # ---------------------------------

    if os.path.exists(chart_path):

        img = XLImage(chart_path)

        img.width = 700

        img.height = 450

        ws1.add_image(
            img,
            "D3"
        )


    # ---------------------------------
    # AUTO WIDTH
    # ---------------------------------

    for sheet in wb.worksheets:

        for column in sheet.columns:

            max_length = 0

            column_letter = (
                get_column_letter(
                    column[0].column
                )
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40
            )


    # ---------------------------------
    # SAVE
    # ---------------------------------

    output_path = os.path.join(
        REPORT_DIR,
        "sales_report.xlsx"
    )

    wb.save(output_path)

    return output_path